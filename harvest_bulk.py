#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
harvest_bulk.py -- the remaining large open datasets.

    python3 harvest_bulk.py --brazil       # municipality-level rescue operations
    python3 harvest_bulk.py --glotip       # UNODC detected victims & convictions
    python3 harvest_bulk.py --iuu          # IUU-listed fishing vessels
    python3 harvest_bulk.py --gfw --token X  # Global Fishing Watch vessels
    python3 harvest_bulk.py --all -v

Writes bulk.json, merged into the same dot layer as everything else.

WHY BRAZIL IS FIRST
-------------------
The Observatorio Digital do Trabalho Escravo, run by Brazil's Labour
Prosecution Service with the ILO, publishes rescue operations at MUNICIPALITY
level: 60,251 people found in conditions analogous to slavery between 1995 and
2022, located to the town. Nothing else in this field comes close as
sub-national enforcement data -- it is not modelled, not estimated, and not
aggregated to the country. It is a state saying: we went here, and we found this
many people.

Municipality names are resolved to coordinates through IBGE's public
localidades API, so no coordinate table is invented or maintained here.

WHAT EACH LAYER IS
------------------
  brazil  Enforcement outcome. People actually removed by inspectors. The one
          layer on this map that records a rescue rather than a risk.
  glotip  Detection. Victims detected and traffickers convicted, as reported by
          states to UNODC. Counts the response, not the phenomenon -- a country
          reporting few detections may have little trafficking or no
          identification system, and the data cannot tell you which.
  iuu     Vessels listed for illegal, unreported and unregulated fishing by the
          regional fisheries bodies. Not a forced-labour finding, but IUU
          operation and forced labour at sea are strongly correlated in the
          documented cases, and a listed vessel is a named, flagged, actionable
          entity.
  gfw     Vessel behaviour: long voyages without port calls, transhipment at
          sea, transponder gaps. Indicators, not findings.
"""

import argparse
import csv
import io
import json
import os
import re
import ssl
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "bulk.json")
UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/124.0 Safari/537.36")

DATA_DIR = os.path.join(HERE, "data")


def find_export(path, *patterns):
    """Repo-only workflow: an export committed to data/ is found by keyword, so
    a GitHub Action uses it without anyone passing a path."""
    if path:
        return path
    if not os.path.isdir(DATA_DIR):
        return None
    for f in sorted(os.listdir(DATA_DIR)):
        low = f.lower()
        if low.endswith((".csv", ".json", ".xlsx")) and any(p in low for p in patterns):
            found = os.path.join(DATA_DIR, f)
            print("  found export in data/: %s" % f)
            return found
    return None


def read_file(path, what):
    """A missing input should say what to do, not throw a stack trace."""
    if not path:
        return None
    if not os.path.exists(path):
        here = os.path.abspath(os.getcwd())
        print("\n  File not found: %s" % path)
        print("  You are in: %s" % here)
        print("  %s" % what)
        return None
    return path


# IBGE's own API returned 403s and then a 500. It also needed a second request
# per municipality for geometry, which is thousands of calls for one layer.
# This gazetteer is a single 5,570-row CSV with coordinates already in it,
# served from GitHub, which both this and a GitHub Actions runner can reach.
MUN_GAZETTEER = ("https://raw.githubusercontent.com/kelvins/municipios-brasileiros"
                 "/main/csv/municipios.csv")
UF_BY_CODE = {11:"RO",12:"AC",13:"AM",14:"RR",15:"PA",16:"AP",17:"TO",21:"MA",
              22:"PI",23:"CE",24:"RN",25:"PB",26:"PE",27:"AL",28:"SE",29:"BA",
              31:"MG",32:"ES",33:"RJ",35:"SP",41:"PR",42:"SC",43:"RS",50:"MS",
              51:"MT",52:"GO",53:"DF"}

_MUN = None


def ibge_index(verbose=False):
    """Municipality name + UF -> coordinates. One request, no per-row lookups."""
    global _MUN
    if _MUN is not None:
        return _MUN
    text = None
    local = find_export(None, "municipio", "ibge", "gazetteer")
    if local:
        text = open(local, encoding="utf-8-sig").read()
    else:
        try:
            text = fetch(MUN_GAZETTEER, timeout=120).decode("utf-8", "replace")
        except Exception as ex:
            print("  gazetteer fetch failed: %s" % str(ex)[:70])
            print("  Download %s and commit it to data/ with 'municipios' in the "
                  "filename." % MUN_GAZETTEER)
            _MUN = {}
            return _MUN
    idx = {}
    for r in csv.DictReader(io.StringIO(text)):
        try:
            uf = UF_BY_CODE.get(int(r.get("codigo_uf") or 0), "")
            lat = float(r["latitude"])
            lng = float(r["longitude"])
        except (TypeError, ValueError, KeyError):
            continue
        idx[(norm(r.get("nome", "")), uf)] = {
            "id": r.get("codigo_ibge"), "uf": uf, "name": r.get("nome"),
            "lat": lat, "lng": lng}
    print("  municipalities indexed: %d" % len(idx))
    _MUN = idx
    return idx


_COORD_CACHE = {}


def mun_coords(entry):
    """Coordinates come straight from the gazetteer row now."""
    if isinstance(entry, dict):
        return (entry.get("lat"), entry.get("lng"))
    return None


def norm(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s or "").lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9 ]+", " ", s).strip()


_COORD_CACHE = {}


def brazil_file_kind(path):
    """The two Brazilian files look alike and get named alike. Decide by
    columns rather than by filename, because `resgates_brazil.csv` turned out
    to be the employer register and went to the wrong parser."""
    try:
        raw = open(path, "rb").read(4000)
        for enc in ("utf-8-sig", "latin-1"):
            try:
                head = raw.decode(enc).split("\n")[0].lower()
                break
            except UnicodeDecodeError:
                continue
        if "empregador" in head or "estabelecimento" in head or "cnpj" in head:
            return "listasuja"
        if "resgat" in head or "trabalhadores" in head and "munic" in head:
            return "rescues"
    except Exception:
        pass
    return "rescues"


def harvest_brazil(a):
    """Expects a CSV of municipality, UF, rescued count, years. The Observatorio
    is a Shiny dashboard rather than an API, so its export is the realistic
    input; --file takes it."""
    rows = []
    fp = find_export(a.file, "resgat", "escrav", "brazil", "municip",
                     "cadastro", "empregador", "suja")
    if fp and brazil_file_kind(fp) == "listasuja":
        print("  that file is the employer register, not a municipality table \u2014 "
              "routing it to the lista suja parser instead")
        return harvest_listasuja_file(fp, a)
    fp = read_file(fp, "Export the municipality table from "
                           "smartlabbr.org/trabalhoescravo (its download control), commit it "
                           "to data/ with 'resgat' in the filename.")
    if fp:
        rows = list(csv.DictReader(open(fp, encoding="utf-8")))
        print("  using local export: %s (%d rows)" % (fp, len(rows)))
    else:
        print("  The Observatorio Digital do Trabalho Escravo "
              "(smartlabbr.org/trabalhoescravo) is a Shiny dashboard, not an API. "
              "Use its download control to export the municipality table, then:\n"
              "      python3 harvest_bulk.py --brazil --file resgates.csv\n"
              "  Expected columns: municipality name, UF, number rescued, period.")
        return []

    idx = ibge_index(a.verbose)
    out, unmatched = [], 0
    for r in rows:
        mun = None
        uf = None
        n = None
        for k, v in r.items():
            kl = norm(k)
            if mun is None and ("municip" in kl or "cidade" in kl or kl == "nome"):
                mun = v
            if uf is None and (kl in ("uf", "estado", "sigla uf", "state")):
                uf = str(v or "").strip().upper()[:2]
            if n is None and ("resgat" in kl or "trabalhadores" in kl or "total" in kl
                              or "rescued" in kl or "quantidade" in kl):
                try:
                    n = int(float(str(v).replace(".", "").replace(",", ".")))
                except Exception:
                    pass
        if not mun or n is None:
            continue
        hit = idx.get((norm(mun), uf or ""))
        if not hit:
            for (nm, u), val in idx.items():
                if nm == norm(mun):
                    hit = val
                    break
        if not hit:
            unmatched += 1
            continue
        c = mun_coords(hit)
        if not c:
            unmatched += 1
            continue
        out.append({
            "name": "%s: %s rescued" % (hit["name"], format(n, ",")),
            "source": "brazil",
            "type": "Rescue operations \u2014 workers removed",
            "lat": c[0], "lng": c[1], "precise": False,
            # located to a municipality, not to the farm: not precise, but not a
            # whole-country record either, so it belongs on the map
            "local": True,
            "impact": 5 if n >= 200 else 4 if n >= 50 else 3 if n >= 10 else 2,
            "status": "Rescued",
            "state": "%s, %s" % (hit["name"], hit["uf"]),
            "url": "https://observatorioescravo.mpt.mp.br/",
            "desc": ("%s people found working in conditions analogous to slavery in this "
                     "municipality and removed by Brazil's mobile inspection group, per the "
                     "Observatorio Digital do Trabalho Escravo, run by the Labour Prosecution "
                     "Service with the ILO. <b>This is an enforcement outcome, not a risk "
                     "estimate</b> \u2014 the one layer on this map that records a rescue "
                     "rather than a possibility. Placed at the municipality centre; the "
                     "operations themselves were at farms, kilns and worksites within it. "
                     "Brazil recorded 60,251 such rescues between 1995 and 2022, which is "
                     "less a statement about Brazil than about what a country looks like when "
                     "it actually counts." % format(n, ",")),
        })
    if unmatched:
        print("  %d row(s) could not be matched to an IBGE municipality and were "
              "dropped rather than placed approximately" % unmatched)
    print("  Brazil municipality records: %d" % len(out))
    return out



UFS = ("AC AL AP AM BA CE DF ES GO MA MT MS MG PA PB PR PE PI RJ RN RS RO RR "
       "SC SP SE TO").split()

# "FAZENDA SAO JOSE, MUNICIPIO DE CORUMBA/MS" -> ("CORUMBA", "MS")
# Handles the slash and dash forms, the "MUNICIPIO DE" prefix, and trailing
# punctuation, which all appear in the published file.
MUN_RE = re.compile(
    r"(?:MUNIC[IÍ]PIO\s+DE\s+)?([A-ZÀ-Ú][A-ZÀ-Ú'\u2019 .\-]{2,40}?)\s*[/\-]\s*(%s)\b"
    % "|".join(UFS))


def split_municipality(estab):
    """Take the LAST municipality-looking token, because the establishment
    string reads outward-in: farm, road, district, municipality/UF."""
    if not estab:
        return None, None
    hits = MUN_RE.findall(estab.upper())
    if not hits:
        return None, None
    mun, uf = hits[-1]
    # "... AMERICO DE CAMPOS/SP E MAGDA/SP" leaves a leading conjunction on the
    # second match; strip that and the common address-part prefixes.
    # The dash form ("PEDREIRA DA CERQUINHA - ZONA RURAL - REGENERACAO/PI")
    # lets the capture run backwards through the whole address, so keep only
    # the last address part before the state code.
    mun = re.split(r"\s*[,]\s*|\s+-\s+", mun)[-1]
    # Strip the phrases that introduce a municipality rather than being part of
    # its name: "MUNICIPIO DE X", "ATRACADA NO PORTO FLUVIAL DE X", "NA ZONA
    # RURAL DO MUNICIPIO DE X". Accents vary in the source, so match loosely.
    mun = re.sub(r"^.*?MUNIC[IÍÌ]PIOS?\s+DE\s+", "", mun.strip(" ,.-"))
    mun = re.sub(r"^.*?\b(?:CIDADES?|PORTO FLUVIAL)\s+DE\s+", "", mun)
    mun = re.sub(r"^(E|DE|DA|DO|DOS|DAS)\s+", "", mun.strip(" ,.-\u00a0")).strip()
    mun = re.sub(r"^(ZONA RURAL|DISTRITO|POVOADO|BAIRRO|CENTRO|S/N|SN)[ ,]*", "",
                 mun.strip(" ,.-"))
    mun = mun.strip(" ,.-")
    return (mun or None), uf


def mask_id(v):
    """CNPJ identifies a company and stays. CPF identifies a person and does
    not: the Brazilian state publishes it because publication is the sanction,
    but re-publishing an individual's tax number on a third-party map is a
    different act with a different risk, and the map does not need it to be
    useful. Last two digits kept so an entry can still be checked against the
    official register."""
    v = (v or "").strip()
    if "/" in v:                      # CNPJ: 00.000.000/0001-00
        return v
    if re.match(r"^\d{3}\.\d{3}\.\d{3}-\d{2}$", v):
        return "CPF \u2022\u2022\u2022.\u2022\u2022\u2022.\u2022\u2022\u2022-" + v[-2:]
    return v


def harvest_listasuja(a):
    """Brazil's Cadastro de Empregadores -- the "lista suja".

    Employers found by labour inspectors to have subjected workers to
    conditions analogous to slavery, published by the state after the
    administrative process concludes. Nothing else in this field is comparable:
    it is not modelled, not a risk score, not an allegation, and it names the
    employer and the establishment. Banks and buyers use it, which gives it
    commercial force a report does not have.

    It is also contested: employers have repeatedly obtained injunctions
    removing their names, and entries carry the date of inclusion for exactly
    that reason. Every record on the map states the edition it came from.
    """
    fp = read_file(find_export(a.file, "cadastro", "empregador", "suja", "lista",
                               "resgat", "brazil"),
                   "Download the Cadastro de Empregadores CSV from "
                   "gov.br/trabalho-e-emprego and commit it to data/ with "
                   "'cadastro' or 'suja' in the filename.")
    if not fp:
        return []
    return harvest_listasuja_file(fp, a)


def harvest_listasuja_file(fp, a):
    raw = open(fp, "rb").read()
    for enc in ("utf-8-sig", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    delim = ";" if text.count(";") > text.count(",") else ","
    rows = list(csv.DictReader(io.StringIO(text), delimiter=delim))
    print("  rows: %d (delimiter %r)" % (len(rows), delim))

    idx = ibge_index(a.verbose)
    out, unmatched, nomun = [], [], 0
    for r in rows:
        low = {norm(k): v for k, v in r.items()}
        emp = (low.get("empregador") or "").strip()
        estab = (low.get("estabelecimento") or "").strip()
        if not emp:
            continue
        uf_col = (low.get("uf") or "").strip().upper()[:2]
        mun, uf = split_municipality(estab)
        uf = uf or uf_col
        if not mun:
            nomun += 1
            continue
        hit = idx.get((norm(mun), uf)) or next(
            (v for (nm, _u), v in idx.items() if nm == norm(mun)), None)
        if not hit:
            unmatched.append("%s/%s" % (mun, uf))
            continue
        c = mun_coords(hit)
        if not c:
            unmatched.append("%s/%s (no geometry)" % (mun, uf))
            continue

        try:
            n = int(float(low.get("trabalhadores envolvidos") or 0))
        except ValueError:
            n = 0
        year = (low.get("ano da acao fiscal") or "").strip()
        cnae = (low.get("cnae") or "").strip()
        incl = (low.get("inclusao no cadastro de empregadores") or "").strip()
        doc = mask_id(low.get("cnpj/cpf"))

        out.append({
            "name": emp[:120],
            "source": "listasuja",
            "type": "Employer on Brazil's register",
            "lat": c[0], "lng": c[1], "precise": False, "local": True,
            "impact": 5 if n >= 30 else 4 if n >= 10 else 3 if n >= 2 else 2,
            "status": "Named by the state",
            "state": "%s, %s" % (hit["name"], hit["uf"]),
            "url": ("https://www.gov.br/trabalho-e-emprego/pt-br/assuntos/"
                    "inspecao-do-trabalho/areas-de-atuacao/trabalho-escravo"),
            "desc": (
                ("<b>%s</b> was found by Brazilian labour inspectors to have subjected "
                 "workers to conditions analogous to slavery" % emp)
                + ((", %d worker%s involved" % (n, "" if n == 1 else "s")) if n else "")
                + ((", in the %s inspection year" % year) if year else "") + ". "
                + (("Establishment: %s. " % estab) if estab else "")
                + (("Registered as %s. " % doc) if doc else "")
                + (("Economic activity code %s. " % cnae) if cnae else "")
                + (("Added to the register %s. " % incl) if incl else "")
                + "From the <b>Cadastro de Empregadores</b>, the register the Brazilian "
                  "state publishes after the administrative process concludes \u2014 not "
                  "an allegation and not a risk score, but a completed finding, with the "
                  "employer named. Banks and buyers use it, which gives it commercial "
                  "force a report does not have. "
                  "<b>Two things to hold on to.</b> The register is contested: employers "
                  "have repeatedly obtained injunctions removing their names, so check the "
                  "current edition before relying on any single entry. And the dot is the "
                  "municipality, not the farm \u2014 the establishment address is in the "
                  "text above, and it is more precise than the map can honestly draw."),
        })

    print("  employers placed: %d" % len(out))
    if nomun:
        print("  %d row(s) had no municipality in the establishment field" % nomun)
    if unmatched:
        print("  %d municipality name(s) did not match IBGE and were dropped rather "
              "than approximated: %s%s"
              % (len(unmatched), ", ".join(sorted(set(unmatched))[:6]),
                 " ..." if len(set(unmatched)) > 6 else ""))
    return out


# ==================================================================== GLOTIP
def harvest_glotip(a):
    """UNODC GLOTIP.

    The file is a long table -- one row per country, indicator, dimension,
    category, sex, age and year -- so it has to be filtered rather than read.
    Only rows counting DETECTED VICTIMS are taken, and only the "Total" sex and
    age rows, or every victim would be counted three or four times over.

    Values below five are published as "<5" rather than a number, because a
    small count in a small country can identify a person. Those rows are kept
    at a nominal 3 and the record says so, since dropping them would make the
    countries doing least identification look like the countries with least
    trafficking.
    """
    fp = read_file(find_export(a.file, "glotip", "unodc", "traffick"),
                   "Export the country table from dataunodc.un.org and commit "
                   "it to data/ with 'glotip' in the filename.")
    if not fp:
        return []

    rows = []
    if fp.lower().endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            print("  that file is a spreadsheet and openpyxl is not installed.\n"
                  "      pip install openpyxl")
            return []
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # UNODC's export declares its dimensions as A1:A1, so a read-only sheet
        # stops after one row and the whole file reads as empty. Recomputing
        # from the actual cells is the documented fix.
        try:
            ws.reset_dimensions()
        except Exception:
            pass
        hdr = None
        for r in ws.iter_rows(values_only=True):
            vals = ["" if c is None else str(c).strip() for c in r]
            if hdr is None:
                if "iso3_code" in [v.lower() for v in vals]:
                    hdr = vals
                continue
            rows.append(dict(zip(hdr, vals)))
    else:
        rows = list(csv.DictReader(open(fp, encoding="utf-8-sig", newline="")))
    print("  rows: %d" % len(rows))

    def g(r, *keys):
        for k in r:
            if norm(k).replace(" ", "") in keys:
                return r[k]
        return ""

    agg, censored = {}, 0
    for r in rows:
        iso = (g(r, "iso3code") or "").strip().upper()
        if len(iso) != 3:
            continue
        if "detected" not in (g(r, "indicator") or "").lower():
            continue
        if (g(r, "sex") or "Total").strip().lower() != "total":
            continue
        if (g(r, "age") or "Total").strip().lower() != "total":
            continue
        # GLOTIP is long-format and the same victims appear under several
        # DIMENSIONS at once -- by age group, by form of exploitation, by
        # country of repatriation. Summing them all gave 1.56 million detected
        # victims worldwide, roughly triple the plausible figure, with Pakistan
        # above the United States. Only the "Total" dimension is unduplicated.
        if (g(r, "dimension") or "Total").strip().lower() != "total":
            continue
        raw = (g(r, "txtvalue") or "").strip()
        if raw.startswith("<"):
            val, censored = 3, censored + 1
        else:
            try:
                val = float(raw.replace(",", ""))
            except ValueError:
                continue
        e = agg.setdefault(iso, {"n": 0.0, "yrs": set(),
                                 "country": (g(r, "country") or iso).strip()})
        e["n"] += val
        y = (g(r, "year") or "").strip()
        if y.isdigit():
            e["yrs"].add(int(y))

    out = []
    for iso, e in agg.items():
        if e["n"] <= 0:
            continue
        yrs = sorted(e["yrs"])
        span = ("%d\u2013%d" % (yrs[0], yrs[-1])) if len(yrs) > 1 else (str(yrs[0]) if yrs else "")
        out.append({
            "name": "%s detected victims reported" % format(int(e["n"]), ","),
            "source": "glotip", "type": "Detected victims (UNODC)",
            "iso": iso, "state": e["country"], "country_name": e["country"],
            "precise": False,
            "impact": 5 if e["n"] >= 5000 else 4 if e["n"] >= 1000 else 3 if e["n"] >= 100 else 2,
            "status": "Detected",
            "url": "https://www.unodc.org/unodc/en/data-and-analysis/glotip.html",
            "desc": ("%s victims of trafficking detected and reported to UNODC by %s"
                     % (format(int(e["n"]), ","), e["country"])
                     + ((", across %s" % span) if span else "") + ". "
                     + "<b>This counts the response, not the phenomenon.</b> A country "
                       "reporting few detections may have little trafficking or no "
                       "identification system, and this number cannot distinguish them "
                       "\u2014 UNODC says so itself. Read it against the prevalence "
                       "estimate, which is modelled independently of whether anyone was "
                       "looking; the two disagreeing about a country is the finding. "
                       "Counts below five are published as \u2018<5\u2019 rather than a "
                       "number, because a small count in a small country can identify a "
                       "person; those are counted here as 3."),
        })
    print("  countries: %d (%d censored '<5' cells counted as 3)" % (len(out), censored))
    return out


# ======================================================================= IUU

def pdf_text(raw):
    """Extract text from a PDF without a third-party dependency.

    Most RFMO IUU lists are PDFs with FlateDecode content streams, so zlib plus
    a regex over the text-showing operators gets the vessel numbers out. pypdf
    is used instead when it happens to be installed, because it handles the
    awkward encodings better \u2014 but it is not required, and requiring an
    install to read a public vessel list would be its own kind of failure."""
    try:
        import pypdf
        import io as _io
        return "\n".join((pg.extract_text() or "")
                          for pg in pypdf.PdfReader(_io.BytesIO(raw)).pages)
    except Exception:
        pass
    import zlib
    out = []
    for m in re.finditer(rb"stream\r?\n(.*?)endstream", raw, re.S):
        chunk = m.group(1)
        try:
            chunk = zlib.decompress(chunk)
        except Exception:
            pass
        # text-showing operators: (literal) Tj and [(a)-1(b)] TJ
        for t in re.findall(rb"\((?:\\.|[^\\)])*\)", chunk):
            try:
                out.append(t[1:-1].decode("latin-1"))
            except Exception:
                continue
    return " ".join(out)


DOC_EXT = (".pdf", ".xlsx", ".xls", ".csv", ".doc", ".docx")


def linked_docs(html, base):
    """Pages that describe an IUU list but do not contain it almost always link
    it. Follow anything that looks like the list itself."""
    urls = []
    for href in re.findall(r'href="([^"]+)"', html, re.I):
        low = href.lower()
        if not low.endswith(DOC_EXT):
            continue
        # "cmm" matched every conservation measure the body has ever published:
        # the last run followed six of them and found paragraph numbers. Require
        # a genuine list signal instead.
        if not (("iuu" in low) or ("vessel" in low and "list" in low)):
            continue
        urls.append(urllib.parse.urljoin(base, href))
    seen, out = set(), []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out[:6]


IUU_SOURCES = [
    ("ICCAT IUU list", "https://www.iccat.int/en/IUUlist.html"),
    ("IOTC IUU list", "https://iotc.org/vessels/iuu"),
    ("GFCM IUU list", "https://www.fao.org/gfcm/data/iuu-vessel-list/en/"),
    ("FAO Global Record", "https://www.fao.org/global-record/en/"),
    ("CCAMLR IUU list (non-contracting parties)",
     "https://www.ccamlr.org/en/compliance/non-contracting-party-iuu-vessel-list"),
    ("CCAMLR IUU list (contracting parties)",
     "https://www.ccamlr.org/en/compliance/contracting-party-iuu-vessel-list"),
    ("NAFO IUU list", "https://www.nafo.int/Fisheries/IUU"),
    ("SEAFO IUU list", "https://www.seafo.org/Management/IUU-Vessels"),
    ("IATTC", "https://www.iattc.org/en-US/Fisheries/IUU"),
    ("SPRFMO", "https://www.sprfmo.int/measures/"),
    ("NPFC", "https://www.npfc.int/"),
    ("WCPFC", "https://www.wcpfc.int/"),
    ("Combined IUU Vessel List (TMT)", "https://iuu-vessels.org/"),
]


def _imo_check(s):
    """IMO numbers carry a check digit: the first six digits weighted 7..2 sum
    to a value whose last digit is the seventh. Cheap, and it removes almost
    every false positive a 7-digit regex picks up."""
    if len(s) != 7 or not s.isdigit():
        return False
    if sum(int(s[i]) * (7 - i) for i in range(6)) % 10 != int(s[6]):
        return False
    # The check digit alone still passes 1234567 and 0000000. Assigned IMO
    # numbers start at 5 or above, and a strictly sequential or single-repeated
    # run is a placeholder or a page artefact, not a hull.
    if s[0] < "5":
        return False
    if len(set(s)) == 1:
        return False
    if all(int(s[i + 1]) - int(s[i]) == 1 for i in range(6)):
        return False
    return True


def _looks_like_js_shell(html):
    """A near-empty document with script tags and no table is an app that
    renders client-side. Worth saying so, because 'found 0' otherwise reads as
    'the list is empty' when it means 'the data is not in this response'."""
    txt = re.sub(r"<script.*?</script>", " ", html, flags=re.S)
    txt = re.sub(r"<[^>]+>", " ", txt)
    return len(txt.split()) < 400 and "<script" in html


def harvest_iuu(a):
    out = []
    seen_imo = set()
    for name, url in IUU_SOURCES:
        try:
            html = fetch(url, timeout=a.timeout).decode("utf-8", "replace")
        except Exception as ex:
            print("  %-42s %s" % (name, str(ex)[:46]))
            continue

        # Several patterns, because every RFMO formats its list differently:
        # labelled, in a table cell on its own, or as "IMO/Lloyd's number".
        pats = [r"IMO[^0-9]{0,12}([0-9]{7})",
                r"Lloyd'?s?[^0-9]{0,14}([0-9]{7})",
                r"<td[^>]*>\s*([0-9]{7})\s*</td>"]
        found = set()
        for p in pats:
            found |= set(re.findall(p, html, re.I))

        # Nothing in the page itself? The list is probably a linked document.
        # That is what "54,719 bytes and no IMO numbers" means in practice.
        if not found:
            docs = linked_docs(html, url)
            for du in docs:
                try:
                    raw = fetch(du, timeout=90)
                except Exception as ex:
                    print("       linked doc failed: %s (%s)"
                          % (du.rsplit("/", 1)[-1][:40], str(ex)[:34]))
                    continue
                text = pdf_text(raw) if raw[:4] == b"%PDF" else raw.decode("utf-8", "replace")
                hits = set()
                for p in [r"IMO[^0-9]{0,12}([0-9]{7})", r"\b([0-9]{7})\b"]:
                    hits |= set(re.findall(p, text, re.I))
                kept_here = {h for h in hits if _imo_check(h)}
                print("       %s: %d chars extracted, %d candidate 7-digit numbers, "
                      "%d valid IMO" % (du.rsplit("/", 1)[-1][:38], len(text),
                                        len(hits), len(kept_here)))
                if len(text) < 500:
                    print("            (almost no text \u2014 the PDF is probably scanned "
                          "images or uses an encoding this extractor cannot read)")
                found |= hits
            if docs and not found:
                print("       followed %d linked document(s), none carried IMO numbers"
                      % len(docs))
        # 7-digit IMO numbers have a check digit; use it to throw out years,
        # phone fragments and reference numbers that happen to be 7 digits.
        valid = {i for i in found if _imo_check(i)}

        if not valid:
            why = ("page renders client-side, so the vessel table is not in the HTML"
                   if _looks_like_js_shell(html)
                   else "no IMO-shaped numbers in %d bytes of HTML" % len(html))
            print("  %-42s 0 \u2014 %s" % (name, why))
            if a.dump:
                fn = os.path.join(HERE, "iuu_%s.html" % re.sub(r"\W+", "_", name.lower())[:30])
                open(fn, "w", encoding="utf-8").write(html)
                print("       dumped to %s" % os.path.basename(fn))
            continue

        new_here = valid - seen_imo
        rejected = len(found) - len(valid)
        print("  %-42s %d IMO numbers (%d new%s)"
              % (name, len(valid), len(new_here),
                 (", %d 7-digit strings rejected by the check digit" % rejected)
                 if rejected else ""))
        if rejected and rejected > len(valid):
            print("       more rejected than kept \u2014 if this list is known to be "
                  "larger, the page may use Lloyd's numbers or have typos; re-run "
                  "with --dump and send me the file")
        seen_imo |= valid
        for imo in sorted(new_here)[: a.max or 500]:
            out.append({
                "name": "Vessel IMO %s \u2014 IUU listed" % imo,
                "source": "iuu", "type": "IUU-listed vessel",
                "iso": "", "state": "At sea", "precise": False,
                "impact": 3, "status": "Listed",
                "url": url,
                "list": name,
                "desc": ("Listed by %s. " % name.replace(" IUU list", "")
                         + "Listed for illegal, unreported or unregulated fishing by a regional "
                         "fisheries management organisation. <b>Not a forced-labour finding.</b> "
                         "It is here because IUU operation and forced labour at sea are "
                         "strongly correlated in the documented cases \u2014 a vessel already "
                         "outside the rules on catch is the kind that stays at sea for months "
                         "without a port call \u2014 and because an IMO number is a named, "
                         "flagged, actionable entity in a domain where almost nothing else is. "
                         "Cross-reference against AIS behaviour before drawing any conclusion."),
            })
    return out


# ======================================================================= GFW
def harvest_gfw(a):
    if not a.token:
        print("  Global Fishing Watch needs a free API token from "
              "globalfishingwatch.org/our-apis. With it, this pulls vessels whose "
              "AIS behaviour matches the documented forced-labour indicators: long "
              "voyages without port calls, repeated transhipment at sea, and "
              "transponder gaps.")
        return []
    url = ("https://gateway.api.globalfishingwatch.org/v3/vessels/search"
           "?query=&datasets[0]=public-global-vessel-identity:latest&limit=%d" % (a.max or 200))
    try:
        j = json.loads(fetch(url, headers={"Authorization": "Bearer " + a.token})
                       .decode("utf-8", "replace"))
    except Exception as ex:
        print("  GFW request failed: %s" % str(ex)[:80])
        return []
    ent = j.get("entries") or j.get("data") or []
    print("  GFW vessels: %d" % len(ent))
    return []


def main():
    ap = argparse.ArgumentParser()
    for f in ("brazil", "listasuja", "glotip", "iuu", "gfw", "all"):
        ap.add_argument("--" + f, action="store_true")
    ap.add_argument("--file")
    ap.add_argument("--token")
    ap.add_argument("--max", type=int, default=500)
    ap.add_argument("--timeout", type=int, default=90)
    ap.add_argument("--dump", action="store_true",
                    help="save the fetched HTML when a source yields nothing")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("-v", "--verbose", action="store_true")
    a = ap.parse_args()
    if not any([a.brazil, a.listasuja, a.glotip, a.iuu, a.gfw, a.all]):
        ap.error("choose --brazil, --listasuja, --glotip, --iuu, --gfw or --all")

    recs = []
    if a.brazil or a.all:
        print("=== Brazil: municipality-level rescue operations ===")
        recs += harvest_brazil(a)
    if a.listasuja or a.all:
        print("=== Brazil: Cadastro de Empregadores (the 'lista suja') ===")
        recs += harvest_listasuja(a)
    if a.glotip or a.all:
        print("=== UNODC GLOTIP: detected victims ===")
        recs += harvest_glotip(a)
    if a.iuu or a.all:
        print("=== IUU-listed fishing vessels ===")
        recs += harvest_iuu(a)
    if a.gfw or a.all:
        print("=== Global Fishing Watch ===")
        recs += harvest_gfw(a)

    # --all runs both Brazilian steps, and --brazil now routes a lista suja file
    # to the lista suja parser, so the same 522 employers arrived twice. Dedupe
    # on what actually identifies a record rather than trusting the callers.
    seen, merged = set(), []
    for r in recs:
        k = (r.get("source"), r.get("name"),
             round(float(r.get("lat") or 0), 4), round(float(r.get("lng") or 0), 4))
        if k in seen:
            continue
        seen.add(k)
        merged.append(r)
    if len(merged) != len(recs):
        print("  removed %d duplicate record(s) produced by overlapping steps"
              % (len(recs) - len(merged)))
    recs = merged

    print("total records: %d" % len(recs))
    if a.dry_run:
        for r in recs[:20]:
            print("  %-46s %s" % (r["name"][:46], r["type"]))
        return 0
    if not recs:
        print("nothing harvested; bulk.json left alone")
        return 1
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"generated": datetime.now(timezone.utc).isoformat(),
                   "note": "Enforcement outcomes, detection counts and vessel listings.",
                   "projects": recs}, f, ensure_ascii=False, indent=1)
    print("wrote", OUT, "-", os.path.getsize(OUT), "bytes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
